# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.contrib.auth.views import logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie
from cliente.models import *
from blog.models import *
# Charts
from django.contrib.auth import get_user_model
from django.http import JsonResponse, HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
import datetime
import time
from dateutil.relativedelta import relativedelta

from django.http import HttpResponseRedirect
from django.core.mail import EmailMessage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from presupuesto.models import PresupuestoForm
from django.views.generic import View
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from .utils import render_to_pdf


def send_email(request):
    if request.method == 'POST':
        form = PresupuestoForm(request.POST)
        if form.is_valid():
            form.save()
            data = form.cleaned_data
            body = data['mensaje'] + '\n' + \
                'Tel.: ' + data['telefono']
            asunto = data['nombre']+ ' ' + '-' + ' ' + data['empresa']
            contacto = data['email']
            email = EmailMessage(asunto, body, 'info@ecolatinapy.com', to=[
                                 'info@ecolatinapy.com'], headers={'Reply-To': contacto})
            email.send()
            return redirect('/')
    else:
        form = PresupuestoForm()

    return render(request, 'home.html', {'form': form})


def blog_list(request):
    if request.method == 'POST':
        form = PresupuestoForm(request.POST)
        if form.is_valid():
            form.save()
            data = form.cleaned_data
            body = data['mensaje'] + '\n' + \
                'Tel.: ' + data['telefono']
            asunto = data['nombre'] + ' ' + '-' + ' ' + data['empresa']
            contacto = data['email']
            email = EmailMessage(asunto, body, 'info@ecolatinapy.com', to=[
                                 'info@ecolatinapy.com'], headers={'Reply-To': contacto})
            email.send()
            return redirect('/')
    else:
        form = PresupuestoForm()

    queryset_list = Post.objects.all().order_by("timestamp")
    paginator = Paginator(queryset_list, 3)

    page = request.GET.get('page')
    try:
        queryset = paginator.page(page)
    except PageNotAnInteger:
        queryset = paginator.page(1)
    except EmptyPage:
        queryset = paginator.page(paginator.num_pages)

    context = {
        "object_list": queryset,
        'form': form,
    }
    return render(request, "blog.html", context)


def blog_detail(request, slug=None):
    if request.method == 'POST':
        form = PresupuestoForm(request.POST)
        if form.is_valid():
            form.save()
            data = form.cleaned_data
            body = data['empresa'] + '\n' + data['nombre'] + \
                '\n' + data['telefono'] + '\n' + data['mensaje']
            asunto = data['email']
            email = EmailMessage(asunto, body, to=['info@ecolatinapy.com'])
            email.send()
            return redirect('/')
    else:
        form = PresupuestoForm()
    instance = get_object_or_404(Post, slug=slug)
    context = {
        "instance": instance,
        'form': form

    }
    return render(request, "blog_detail.html", context)


@ensure_csrf_cookie
def authentication(request):
    if request.user.is_authenticated():
        return HttpResponseRedirect(reverse('panel'))
    else:
        if request.method == 'POST':
            action = request.POST.get('action', None)
            username = request.POST.get('username', None)
            password = request.POST.get('password', None)

            user = authenticate(username=username, password=password)
            login(request, user)
            return HttpResponseRedirect(reverse('panel'))
        return render(request, 'login.html', {})


@login_required(None, 'login', '/login/')
def logout_view(request):
    logout(request)
    return redirect("/")


@login_required(None, 'login', '/login/')
def index_view(request):
    propiedades = Propiedad.objects.filter(cliente__usuario=request.user.id)
    totalPropiedades = propiedades.count()
    cebos = Estacion.objects.filter(
        sector__propiedad__cliente__usuario=request.user.id)
    controles = Control.objects.filter(
        estacion__sector__propiedad__cliente__usuario=request.user.id)
    controles_recientes = controles[:3]
    capturados = controles.filter(e_capturado=True)
    totalPropiedades = propiedades.count()
    totalCebos = cebos.count()
    totalCapturados = capturados.count()
    return render(request, 'index.html', {
        'propiedades': propiedades,
        'totalPropiedades': totalPropiedades,
        'totalCebos': totalCebos,
        'totalCapturados': totalCapturados,
        'controles_recientes': controles_recientes,
    })

@login_required(None, 'login', '/login/')
def property_view(request):
    propiedades = Propiedad.objects.select_related(
        "cliente").filter(cliente__usuario=request.user.id)
    return render(request, 'propiedad.html', {
        'propiedades': propiedades,
    })


@login_required(None, 'login', '/login/')
def sector_view(request, id=None):
    sectores = Sector.objects.filter(propiedad=id)
    propiedad = Propiedad.objects.get(pk=id)
    return render(request, 'sector.html', {
        'sectores': sectores,
        'propiedad': propiedad
    })


@login_required(None, 'login', '/login/')
def station_view(request, id=None):
    estaciones = Estacion.objects.filter(sector=id)
    sector = Sector.objects.get(pk=id)
    return render(request, 'station.html', {
        'sector': sector,
        'estaciones': estaciones
    })


@login_required(None, 'login', '/login/')
def control_view(request, id=None):
    controles = Control.objects.filter(estacion=id)
    estacion = Estacion.objects.get(pk=id)
    return render(request, 'control.html', {
        'estacion': estacion,
        'controles': controles
    })


@login_required(None, 'login', '/login/')
def all_control_view(request):
    controles = Control.objects.filter(
        estacion__sector__propiedad__cliente__usuario=request.user.id)
    if 'daterange' in request.GET:
            date_range = request.GET.get('daterange')
            f_inicio = date_range[:10]
            f_fin = date_range[13:24]
            inicio = datetime.datetime.strptime(f_inicio, '%m/%d/%Y')
            fin = datetime.datetime.strptime(f_fin, '%m/%d/%Y')
            controles = Control.objects.filter(
                fecha__range=[inicio, fin]).order_by('-fecha')
            filtro_msg = "Filtro de fechas entre el " + \
                str(f_inicio)+" y el "+str(f_fin)
            return render(request, 'all_control.html', {'controles': controles, 'filtro_msg': filtro_msg, 'date_range': date_range})
    return render(request, 'all_control.html', {
        'controles': controles,
    })


def ReporteExcel(request):

    wb = Workbook()
    #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE'
    #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    #Creamos los encabezados desde la celda B3 hasta la E3
    # ws['A3'] = '#'
    ws['B3'] = 'SECTOR'
    ws['C3'] = 'PROPIEDAD'
    ws['D3'] = 'FECHA'
    ws['E3'] = 'INICIO'
    ws['F3'] = 'FIN'
    ws['G3'] = 'REPOSICIÓN'
    ws['H3'] = 'INCIDENCIA'
    ws['I3'] = 'OBSERVACIONES'
    cont = 4
    
    if 'daterange' in request.GET:
        date_range = request.GET.get('daterange')
        print(date_range)
        f_inicio = date_range[:10]
        f_fin = date_range[13:24]
        inicio = datetime.datetime.strptime(f_inicio, '%m/%d/%Y')
        fin = datetime.datetime.strptime(f_fin, '%m/%d/%Y')
        controles = Control.objects.filter(
            fecha__range=[inicio, fin], estacion__sector__propiedad__cliente__usuario=request.user.id).order_by('-fecha')
    else:
        print('else')
        controles = Control.objects.filter(
            estacion__sector__propiedad__cliente__usuario=request.user.id)

    #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for control in controles:
        # ws.cell(row=cont, column=1).value = control.estacion
        ws.cell(row=cont, column=2).value = '%s' % (control.estacion.sector)
        ws.cell(row=cont, column=3).value = '%s' % (control.estacion.sector.propiedad)
        ws.cell(row=cont, column=4).value = '%s' % (control.fecha)
        ws.cell(row=cont, column=5).value = '%s' % (control.inicio)
        ws.cell(row=cont, column=6).value = '%s' % (control.fin)
        if control.a_reposicion or control.a_reemplazo:
            ws.cell(row=cont, column=7).value = 'SI'
        else:
            ws.cell(row=cont, column=7).value = 'NO'
        if control.e_ausencia or control.e_consumido or control.e_deteriorado or control.e_pa_deteriorada or control.e_no_acceso:
            ws.cell(row=cont, column=8).value = 'SI'
        else:
            ws.cell(row=cont, column=8).value = 'NO'

        cont = cont + 1
    #Establecemos el nombre del archivo
    nombre_archivo = "ReportePersonasExcel.xlsx"
    #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def GeneratePdf(request):
    if 'daterange' in request.GET:
        date_range = request.GET.get('daterange')
        f_inicio = date_range[:10]
        f_fin = date_range[13:24]
        inicio = datetime.datetime.strptime(f_inicio, '%m/%d/%Y')
        fin = datetime.datetime.strptime(f_fin, '%m/%d/%Y')
        controles = Control.objects.filter(
            fecha__range=[inicio, fin], estacion__sector__propiedad__cliente__usuario=request.user.id).order_by('-fecha')
    else:
        controles = Control.objects.filter(
            estacion__sector__propiedad__cliente__usuario=request.user.id)
    cliente = Cliente.objects.get(usuario=request.user.id)
    data = {
        'controles': controles,
        'cliente': cliente
    }
    pdf = render_to_pdf('pdf/reporte_control.html', data)
    return HttpResponse(pdf, content_type='application/pdf')

class ChartData(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, format=None):
        z = datetime.datetime.now()
        z_mes = z.month
        label_z_mes = z.strftime("%B-%Y")
        cantidad_z_mes = Control.objects.filter(fecha__month=z_mes).count()

        y = z - relativedelta(months=1)
        y_mes = y.month
        label_y_mes = y.strftime("%B-%Y")
        cantidad_y_mes = Control.objects.filter(fecha__month=y_mes).count()

        x = y - relativedelta(months=1)
        x_mes = x.month
        label_x_mes = x.strftime("%B-%Y")
        cantidad_x_mes = Control.objects.filter(fecha__month=x_mes).count()

        w = x - relativedelta(months=1)
        w_mes = w.month
        label_w_mes = w.strftime("%B-%Y")
        cantidad_w_mes = Control.objects.filter(fecha__month=w_mes).count()

        v = w - relativedelta(months=1)
        v_mes = v.month
        label_v_mes = v.strftime("%B-%Y")
        cantidad_v_mes = Control.objects.filter(fecha__month=v_mes).count()

        u = v - relativedelta(months=1)
        u_mes = u.month
        label_u_mes = u.strftime("%B-%Y")
        cantidad_u_mes = Control.objects.filter(fecha__month=u_mes).count()

        labels = [label_u_mes, label_v_mes, label_w_mes,
                  label_x_mes, label_y_mes, label_z_mes]
        default_items = [cantidad_u_mes, cantidad_v_mes, cantidad_w_mes,
                         cantidad_x_mes, cantidad_y_mes, cantidad_z_mes]
        data = {
            "labels": labels,
            "default": default_items,
        }
        return Response(data)
